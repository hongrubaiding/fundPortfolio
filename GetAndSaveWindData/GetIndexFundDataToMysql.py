# -- coding: utf-8 --

'''
    将ifind的数据导入到本地数据库,并从数据库返回结果
'''

import pandas as pd
import os
import xlrd
from openpyxl import load_workbook
import numpy as np
from GetAndSaveWindData.GetDataToMysql import GetDataToMysql
from datetime import datetime
import mylog as mylog


class GetIndexFundDataToMysql:
    def __init__(self):
        self.target_path = "D:\\工作文件\\主题基金\\"
        self.mysql_name_dic = {"证券代码": 'fund_code', '证券名称': "fund_name", "基金简称": "fund_simple_name",
                               "是否指数基金": "is_index_fund", "基金管理人": "fund_company", "管理费率百分": "manage_fee",
                               "托管费率百分": "tg_fee", "跟踪指数同花顺代码": "ths_bench_code", '基金成立日': "establish_date",
                               "上市日期": 'etf_pub_date', "基金经理（现任）": 'manager_name', "是否分级基金": "is_graded_fund",
                               "主题名称": 'topic_name'}
        self.logger = mylog.set_log()

    def get_data_to_mysql(self):
        GetDataToMysqlDemo = GetDataToMysql()
        total_file = os.listdir(self.target_path)
        name_dic = {"管理费率\n[单位]%": "管理费率百分", "托管费率\n[单位]%": "托管费率百分"}
        for file_name in total_file:
            if file_name.find('同花顺主题基金') != -1:
                work_book = load_workbook(self.target_path + file_name)
                all_sheets = work_book.sheetnames
                if all_sheets:
                    for sheet_name in all_sheets:
                        self.logger.info("读取%s主题基金..."%sheet_name)
                        temp_df = pd.read_excel(self.target_path + file_name, sheet_name=sheet_name).iloc[:-2]
                        temp_df.rename(columns=name_dic, inplace=True)
                        temp_df.replace({'上市日期': '--'}, np.nan, inplace=True)
                        temp_df.replace({'是否分级基金': '--'}, '否', inplace=True)
                        temp_df.replace({'跟踪指数同花顺代码': '--'}, np.nan, inplace=True)
                        temp_df['主题名称'] = sheet_name
                        temp_df.rename(columns=self.mysql_name_dic, inplace=True)
                        temp_df['record_time']= datetime.today().strftime("%Y-%m-%d")
                        GetDataToMysqlDemo.GetMain(temp_df,tableName='ths_topic_fund')
                        self.logger.info("%s主题基金存储数据库成功！"%sheet_name)


if __name__ == '__main__':
    GetIndexFundDataToMysqlDemo = GetIndexFundDataToMysql()
    GetIndexFundDataToMysqlDemo.get_data_to_mysql()
